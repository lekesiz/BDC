"""
Comprehensive Reporting Service

This module provides a full-featured reporting service with support for:
- Multiple output formats (PDF, Excel, CSV)
- Template-based report generation
- Scheduled report generation
- Email delivery
- Custom report builder
- Report caching and storage
- History tracking
- Data aggregation and analysis
- Charts and visualizations
- Multi-language support
"""

import asyncio
import json
import os
import smtplib
import tempfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
from uuid import uuid4

import pandas as pd
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from jinja2 import Environment, FileSystemLoader, Template
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
import hashlib
from functools import lru_cache
import aiofiles
import aiosmtplib
from email.message import EmailMessage
import logging
from babel import Locale
from babel.support import Translations
import gettext

logger = logging.getLogger(__name__)


class ReportFormat(Enum):
    """Supported report output formats"""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    HTML = "html"
    JSON = "json"


class ChartType(Enum):
    """Supported chart types"""
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    SCATTER = "scatter"
    HEATMAP = "heatmap"
    HISTOGRAM = "histogram"


class ReportStatus(Enum):
    """Report generation status"""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    DELIVERED = "delivered"


@dataclass
class ChartConfig:
    """Configuration for chart generation"""
    chart_type: ChartType
    data_columns: List[str]
    x_label: Optional[str] = None
    y_label: Optional[str] = None
    title: Optional[str] = None
    colors: Optional[List[str]] = None
    width: int = 8
    height: int = 6


@dataclass
class ReportTemplate:
    """Report template definition"""
    id: str = field(default_factory=lambda: str(uuid4()))
    name: str = ""
    description: str = ""
    template_type: str = "jinja2"  # jinja2, custom
    template_content: str = ""
    template_file: Optional[str] = None
    data_source: Optional[str] = None  # SQL query, API endpoint, etc.
    parameters: Dict[str, Any] = field(default_factory=dict)
    charts: List[ChartConfig] = field(default_factory=list)
    styles: Dict[str, Any] = field(default_factory=dict)
    language: str = "en"
    created_at: datetime = field(default_factory=datetime.utcnow)
    updated_at: datetime = field(default_factory=datetime.utcnow)
    tags: List[str] = field(default_factory=list)


@dataclass
class ReportSchedule:
    """Report scheduling configuration"""
    id: str = field(default_factory=lambda: str(uuid4()))
    report_template_id: str = ""
    cron_expression: str = ""  # e.g., "0 9 * * 1" for every Monday at 9 AM
    timezone: str = "UTC"
    enabled: bool = True
    recipients: List[str] = field(default_factory=list)
    output_formats: List[ReportFormat] = field(default_factory=list)
    parameters: Dict[str, Any] = field(default_factory=dict)
    retention_days: int = 30
    created_at: datetime = field(default_factory=datetime.utcnow)
    last_run: Optional[datetime] = None
    next_run: Optional[datetime] = None


@dataclass
class ReportMetadata:
    """Report generation metadata and history"""
    id: str = field(default_factory=lambda: str(uuid4()))
    template_id: Optional[str] = None
    template_name: str = ""
    generated_at: datetime = field(default_factory=datetime.utcnow)
    generated_by: str = ""
    format: ReportFormat = ReportFormat.PDF
    file_path: str = ""
    file_size: int = 0
    parameters: Dict[str, Any] = field(default_factory=dict)
    status: ReportStatus = ReportStatus.PENDING
    error_message: Optional[str] = None
    generation_time_ms: int = 0
    delivered_to: List[str] = field(default_factory=list)
    cache_key: Optional[str] = None
    expires_at: Optional[datetime] = None


class ReportService:
    """Comprehensive report generation and management service"""
    
    def __init__(
        self,
        storage_path: str = "/tmp/reports",
        template_path: str = "/tmp/report_templates",
        cache_enabled: bool = True,
        cache_ttl_hours: int = 24,
        smtp_config: Optional[Dict[str, Any]] = None,
        locale_path: str = "/tmp/locales"
    ):
        self.storage_path = Path(storage_path)
        self.template_path = Path(template_path)
        self.cache_enabled = cache_enabled
        self.cache_ttl = timedelta(hours=cache_ttl_hours)
        self.smtp_config = smtp_config or {}
        self.locale_path = Path(locale_path)
        
        # Create directories
        self.storage_path.mkdir(parents=True, exist_ok=True)
        self.template_path.mkdir(parents=True, exist_ok=True)
        self.locale_path.mkdir(parents=True, exist_ok=True)
        
        # Initialize components
        self.scheduler = AsyncIOScheduler()
        self.templates: Dict[str, ReportTemplate] = {}
        self.schedules: Dict[str, ReportSchedule] = {}
        self.report_history: List[ReportMetadata] = []
        
        # Jinja2 environment
        self.jinja_env = Environment(
            loader=FileSystemLoader(str(self.template_path)),
            autoescape=True
        )
        
        # Set up matplotlib
        plt.style.use('seaborn-v0_8-darkgrid')
        
        # Translation support
        self.translations: Dict[str, Translations] = {}
        
    async def initialize(self):
        """Initialize the report service"""
        self.scheduler.start()
        await self._load_templates()
        await self._load_schedules()
        await self._setup_scheduled_reports()
        
    async def shutdown(self):
        """Shutdown the report service"""
        self.scheduler.shutdown()
        
    # Template Management
    
    async def create_template(self, template: ReportTemplate) -> ReportTemplate:
        """Create a new report template"""
        template.created_at = datetime.utcnow()
        template.updated_at = datetime.utcnow()
        
        # Save template file if content provided
        if template.template_content and not template.template_file:
            template_filename = f"{template.id}.jinja2"
            template_filepath = self.template_path / template_filename
            async with aiofiles.open(template_filepath, 'w') as f:
                await f.write(template.template_content)
            template.template_file = template_filename
            
        self.templates[template.id] = template
        await self._save_templates()
        return template
        
    async def update_template(self, template_id: str, updates: Dict[str, Any]) -> Optional[ReportTemplate]:
        """Update an existing template"""
        if template_id not in self.templates:
            return None
            
        template = self.templates[template_id]
        for key, value in updates.items():
            if hasattr(template, key):
                setattr(template, key, value)
                
        template.updated_at = datetime.utcnow()
        
        # Update template file if content changed
        if 'template_content' in updates and template.template_file:
            template_filepath = self.template_path / template.template_file
            async with aiofiles.open(template_filepath, 'w') as f:
                await f.write(updates['template_content'])
                
        await self._save_templates()
        return template
        
    async def delete_template(self, template_id: str) -> bool:
        """Delete a template"""
        if template_id not in self.templates:
            return False
            
        template = self.templates[template_id]
        
        # Remove template file
        if template.template_file:
            template_filepath = self.template_path / template.template_file
            if template_filepath.exists():
                template_filepath.unlink()
                
        del self.templates[template_id]
        await self._save_templates()
        return True
        
    async def get_template(self, template_id: str) -> Optional[ReportTemplate]:
        """Get a template by ID"""
        return self.templates.get(template_id)
        
    async def list_templates(self, tags: Optional[List[str]] = None) -> List[ReportTemplate]:
        """List all templates, optionally filtered by tags"""
        templates = list(self.templates.values())
        
        if tags:
            templates = [
                t for t in templates
                if any(tag in t.tags for tag in tags)
            ]
            
        return templates
        
    # Report Generation
    
    async def generate_report(
        self,
        template_id: Optional[str] = None,
        data: Optional[Union[pd.DataFrame, Dict[str, Any], List[Dict[str, Any]]]] = None,
        parameters: Optional[Dict[str, Any]] = None,
        output_format: ReportFormat = ReportFormat.PDF,
        language: str = "en",
        custom_template: Optional[str] = None
    ) -> ReportMetadata:
        """Generate a report using a template or custom content"""
        start_time = datetime.utcnow()
        metadata = ReportMetadata(
            template_id=template_id,
            format=output_format,
            parameters=parameters or {},
            status=ReportStatus.PROCESSING
        )
        
        try:
            # Get template
            template = None
            if template_id:
                template = await self.get_template(template_id)
                if not template:
                    raise ValueError(f"Template {template_id} not found")
                metadata.template_name = template.name
                
            # Check cache
            if self.cache_enabled and template_id:
                cache_key = self._generate_cache_key(template_id, parameters, output_format)
                cached_report = await self._get_cached_report(cache_key)
                if cached_report:
                    return cached_report
                    
            # Prepare data
            if data is None and template and template.data_source:
                data = await self._fetch_data(template.data_source, parameters)
                
            # Convert data to DataFrame if needed
            if isinstance(data, list):
                data = pd.DataFrame(data)
            elif isinstance(data, dict) and not isinstance(data, pd.DataFrame):
                data = pd.DataFrame([data])
                
            # Generate report content
            if template:
                content = await self._render_template(template, data, parameters, language)
            elif custom_template:
                content = await self._render_custom_template(custom_template, data, parameters)
            else:
                content = data
                
            # Generate charts if specified
            charts = []
            if template and template.charts:
                charts = await self._generate_charts(template.charts, data)
                
            # Generate output file
            output_file = await self._generate_output_file(
                content, 
                output_format, 
                charts,
                template.styles if template else {}
            )
            
            # Save report
            report_filename = f"report_{metadata.id}.{output_format.value}"
            report_path = self.storage_path / report_filename
            
            if isinstance(output_file, bytes):
                async with aiofiles.open(report_path, 'wb') as f:
                    await f.write(output_file)
            else:
                async with aiofiles.open(report_path, 'w') as f:
                    await f.write(output_file)
                    
            # Update metadata
            metadata.file_path = str(report_path)
            metadata.file_size = report_path.stat().st_size
            metadata.status = ReportStatus.COMPLETED
            metadata.generation_time_ms = int((datetime.utcnow() - start_time).total_seconds() * 1000)
            
            # Cache report if enabled
            if self.cache_enabled and template_id:
                metadata.cache_key = cache_key
                metadata.expires_at = datetime.utcnow() + self.cache_ttl
                await self._cache_report(metadata)
                
            # Add to history
            self.report_history.append(metadata)
            
            return metadata
            
        except Exception as e:
            metadata.status = ReportStatus.FAILED
            metadata.error_message = str(e)
            self.report_history.append(metadata)
            logger.error(f"Report generation failed: {e}")
            raise
            
    async def _render_template(
        self,
        template: ReportTemplate,
        data: pd.DataFrame,
        parameters: Dict[str, Any],
        language: str
    ) -> str:
        """Render a template with data"""
        # Load template
        if template.template_file:
            jinja_template = self.jinja_env.get_template(template.template_file)
        else:
            jinja_template = Template(template.template_content)
            
        # Prepare context
        context = {
            'data': data.to_dict('records') if isinstance(data, pd.DataFrame) else data,
            'parameters': parameters,
            'template': template,
            'generated_at': datetime.utcnow(),
            'language': language,
        }
        
        # Add translation function
        if language != 'en':
            translations = await self._get_translations(language)
            context['_'] = translations.gettext
            context['ngettext'] = translations.ngettext
        else:
            context['_'] = lambda x: x
            context['ngettext'] = lambda s, p, n: s if n == 1 else p
            
        # Add data aggregation functions
        if isinstance(data, pd.DataFrame):
            context.update({
                'sum': lambda col: data[col].sum(),
                'mean': lambda col: data[col].mean(),
                'count': lambda col: data[col].count(),
                'min': lambda col: data[col].min(),
                'max': lambda col: data[col].max(),
                'groupby': lambda col: data.groupby(col).size().to_dict(),
            })
            
        return jinja_template.render(**context)
        
    async def _render_custom_template(
        self,
        template_content: str,
        data: Any,
        parameters: Dict[str, Any]
    ) -> str:
        """Render a custom template"""
        template = Template(template_content)
        context = {
            'data': data.to_dict('records') if isinstance(data, pd.DataFrame) else data,
            'parameters': parameters,
            'generated_at': datetime.utcnow(),
        }
        return template.render(**context)
        
    async def _generate_charts(
        self,
        chart_configs: List[ChartConfig],
        data: pd.DataFrame
    ) -> List[BytesIO]:
        """Generate charts from data"""
        charts = []
        
        for config in chart_configs:
            plt.figure(figsize=(config.width, config.height))
            
            if config.chart_type == ChartType.BAR:
                if len(config.data_columns) == 2:
                    plt.bar(data[config.data_columns[0]], data[config.data_columns[1]])
                else:
                    data[config.data_columns].plot(kind='bar')
                    
            elif config.chart_type == ChartType.LINE:
                if len(config.data_columns) == 2:
                    plt.plot(data[config.data_columns[0]], data[config.data_columns[1]])
                else:
                    data[config.data_columns].plot(kind='line')
                    
            elif config.chart_type == ChartType.PIE:
                if len(config.data_columns) == 2:
                    plt.pie(data[config.data_columns[1]], labels=data[config.data_columns[0]])
                else:
                    data[config.data_columns[0]].value_counts().plot(kind='pie')
                    
            elif config.chart_type == ChartType.SCATTER:
                plt.scatter(data[config.data_columns[0]], data[config.data_columns[1]])
                
            elif config.chart_type == ChartType.HEATMAP:
                pivot_data = data.pivot_table(
                    values=config.data_columns[2] if len(config.data_columns) > 2 else None,
                    index=config.data_columns[0],
                    columns=config.data_columns[1]
                )
                sns.heatmap(pivot_data, annot=True, cmap='coolwarm')
                
            elif config.chart_type == ChartType.HISTOGRAM:
                plt.hist(data[config.data_columns[0]], bins=20)
                
            # Add labels and title
            if config.title:
                plt.title(config.title)
            if config.x_label:
                plt.xlabel(config.x_label)
            if config.y_label:
                plt.ylabel(config.y_label)
                
            # Save to BytesIO
            chart_buffer = BytesIO()
            plt.savefig(chart_buffer, format='png', bbox_inches='tight')
            chart_buffer.seek(0)
            charts.append(chart_buffer)
            plt.close()
            
        return charts
        
    async def _generate_output_file(
        self,
        content: Union[str, pd.DataFrame],
        output_format: ReportFormat,
        charts: List[BytesIO],
        styles: Dict[str, Any]
    ) -> Union[bytes, str]:
        """Generate output file in specified format"""
        if output_format == ReportFormat.PDF:
            return await self._generate_pdf(content, charts, styles)
        elif output_format == ReportFormat.EXCEL:
            return await self._generate_excel(content, charts, styles)
        elif output_format == ReportFormat.CSV:
            return await self._generate_csv(content)
        elif output_format == ReportFormat.HTML:
            return await self._generate_html(content, charts, styles)
        elif output_format == ReportFormat.JSON:
            return await self._generate_json(content)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
            
    async def _generate_pdf(
        self,
        content: Union[str, pd.DataFrame],
        charts: List[BytesIO],
        styles: Dict[str, Any]
    ) -> bytes:
        """Generate PDF report"""
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Add styles
        styles_sheet = getSampleStyleSheet()
        
        # Add custom styles
        if 'title' in styles:
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles_sheet['Title'],
                **styles['title']
            )
        else:
            title_style = styles_sheet['Title']
            
        # Add content
        if isinstance(content, str):
            # HTML content
            para = Paragraph(content, styles_sheet['Normal'])
            elements.append(para)
        elif isinstance(content, pd.DataFrame):
            # DataFrame as table
            table_data = [content.columns.tolist()] + content.values.tolist()
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
        # Add charts
        for chart in charts:
            elements.append(Spacer(1, 12))
            img = RLImage(chart, width=400, height=300)
            elements.append(img)
            
        # Build PDF
        doc.build(elements)
        
        buffer.seek(0)
        return buffer.read()
        
    async def _generate_excel(
        self,
        content: Union[str, pd.DataFrame],
        charts: List[BytesIO],
        styles: Dict[str, Any]
    ) -> bytes:
        """Generate Excel report"""
        buffer = BytesIO()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Add content
        if isinstance(content, pd.DataFrame):
            # Add headers
            for col_num, header in enumerate(content.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                
            # Add data
            for row_num, row_data in enumerate(content.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                    
            # Add charts
            if charts:
                chart_sheet = wb.create_sheet("Charts")
                for i, chart in enumerate(charts):
                    img = XLImage(chart)
                    chart_sheet.add_image(img, f'A{i*20 + 1}')
                    
        elif isinstance(content, str):
            ws.cell(row=1, column=1, value=content)
            
        # Apply styles
        if 'column_width' in styles:
            for col, width in styles['column_width'].items():
                ws.column_dimensions[col].width = width
                
        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
        
    async def _generate_csv(self, content: Union[str, pd.DataFrame]) -> str:
        """Generate CSV report"""
        if isinstance(content, pd.DataFrame):
            return content.to_csv(index=False)
        else:
            return str(content)
            
    async def _generate_html(
        self,
        content: Union[str, pd.DataFrame],
        charts: List[BytesIO],
        styles: Dict[str, Any]
    ) -> str:
        """Generate HTML report"""
        html_parts = ['<html><head><style>']
        
        # Add default styles
        html_parts.append("""
            body { font-family: Arial, sans-serif; margin: 40px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #4CAF50; color: white; }
            tr:nth-child(even) { background-color: #f2f2f2; }
            .chart { margin: 20px 0; text-align: center; }
        """)
        
        # Add custom styles
        if 'css' in styles:
            html_parts.append(styles['css'])
            
        html_parts.append('</style></head><body>')
        
        # Add content
        if isinstance(content, pd.DataFrame):
            html_parts.append(content.to_html(index=False, classes='report-table'))
        else:
            html_parts.append(content)
            
        # Add charts as base64 images
        for i, chart in enumerate(charts):
            import base64
            chart_data = base64.b64encode(chart.read()).decode()
            html_parts.append(f'<div class="chart"><img src="data:image/png;base64,{chart_data}" /></div>')
            
        html_parts.append('</body></html>')
        
        return ''.join(html_parts)
        
    async def _generate_json(self, content: Union[str, pd.DataFrame]) -> str:
        """Generate JSON report"""
        if isinstance(content, pd.DataFrame):
            return content.to_json(orient='records', date_format='iso')
        else:
            return json.dumps({'content': content}, default=str)
            
    # Scheduling
    
    async def create_schedule(self, schedule: ReportSchedule) -> ReportSchedule:
        """Create a new report schedule"""
        schedule.created_at = datetime.utcnow()
        
        # Validate cron expression
        try:
            trigger = CronTrigger.from_crontab(schedule.cron_expression)
            schedule.next_run = trigger.get_next_fire_time(None, datetime.utcnow())
        except Exception as e:
            raise ValueError(f"Invalid cron expression: {e}")
            
        self.schedules[schedule.id] = schedule
        
        # Add to scheduler if enabled
        if schedule.enabled:
            await self._add_scheduled_job(schedule)
            
        await self._save_schedules()
        return schedule
        
    async def update_schedule(self, schedule_id: str, updates: Dict[str, Any]) -> Optional[ReportSchedule]:
        """Update a schedule"""
        if schedule_id not in self.schedules:
            return None
            
        schedule = self.schedules[schedule_id]
        
        # Remove old job if exists
        if self.scheduler.get_job(schedule_id):
            self.scheduler.remove_job(schedule_id)
            
        # Update schedule
        for key, value in updates.items():
            if hasattr(schedule, key):
                setattr(schedule, key, value)
                
        # Validate and update next run if cron changed
        if 'cron_expression' in updates:
            try:
                trigger = CronTrigger.from_crontab(schedule.cron_expression)
                schedule.next_run = trigger.get_next_fire_time(None, datetime.utcnow())
            except Exception as e:
                raise ValueError(f"Invalid cron expression: {e}")
                
        # Re-add job if enabled
        if schedule.enabled:
            await self._add_scheduled_job(schedule)
            
        await self._save_schedules()
        return schedule
        
    async def delete_schedule(self, schedule_id: str) -> bool:
        """Delete a schedule"""
        if schedule_id not in self.schedules:
            return False
            
        # Remove from scheduler
        if self.scheduler.get_job(schedule_id):
            self.scheduler.remove_job(schedule_id)
            
        del self.schedules[schedule_id]
        await self._save_schedules()
        return True
        
    async def _add_scheduled_job(self, schedule: ReportSchedule):
        """Add a job to the scheduler"""
        self.scheduler.add_job(
            self._run_scheduled_report,
            CronTrigger.from_crontab(schedule.cron_expression),
            id=schedule.id,
            args=[schedule.id],
            timezone=schedule.timezone,
            replace_existing=True
        )
        
    async def _run_scheduled_report(self, schedule_id: str):
        """Run a scheduled report"""
        schedule = self.schedules.get(schedule_id)
        if not schedule or not schedule.enabled:
            return
            
        try:
            # Update last run time
            schedule.last_run = datetime.utcnow()
            
            # Generate reports in all requested formats
            reports = []
            for format in schedule.output_formats:
                metadata = await self.generate_report(
                    template_id=schedule.report_template_id,
                    parameters=schedule.parameters,
                    output_format=format
                )
                reports.append(metadata)
                
            # Send reports to recipients
            if schedule.recipients:
                await self._deliver_reports(reports, schedule.recipients)
                
            # Update next run time
            trigger = CronTrigger.from_crontab(schedule.cron_expression)
            schedule.next_run = trigger.get_next_fire_time(None, datetime.utcnow())
            
            # Clean up old reports
            await self._cleanup_old_reports(schedule.retention_days)
            
        except Exception as e:
            logger.error(f"Failed to run scheduled report {schedule_id}: {e}")
            
    # Email Delivery
    
    async def deliver_report(
        self,
        report_metadata: ReportMetadata,
        recipients: List[str],
        subject: Optional[str] = None,
        body: Optional[str] = None
    ) -> bool:
        """Deliver a report via email"""
        if not self.smtp_config:
            logger.error("SMTP configuration not provided")
            return False
            
        try:
            # Create message
            msg = EmailMessage()
            msg['From'] = self.smtp_config.get('from_email', 'reports@example.com')
            msg['To'] = ', '.join(recipients)
            msg['Subject'] = subject or f"Report: {report_metadata.template_name or 'Custom Report'}"
            
            # Set body
            if body:
                msg.set_content(body)
            else:
                msg.set_content(f"""
                Please find attached your requested report.
                
                Report Details:
                - Generated: {report_metadata.generated_at}
                - Format: {report_metadata.format.value}
                - Size: {report_metadata.file_size / 1024:.2f} KB
                """)
                
            # Attach report file
            with open(report_metadata.file_path, 'rb') as f:
                file_data = f.read()
                filename = os.path.basename(report_metadata.file_path)
                msg.add_attachment(
                    file_data,
                    maintype='application',
                    subtype='octet-stream',
                    filename=filename
                )
                
            # Send email
            async with aiosmtplib.SMTP(
                hostname=self.smtp_config['host'],
                port=self.smtp_config.get('port', 587),
                use_tls=self.smtp_config.get('use_tls', True)
            ) as smtp:
                if self.smtp_config.get('username'):
                    await smtp.login(
                        self.smtp_config['username'],
                        self.smtp_config['password']
                    )
                await smtp.send_message(msg)
                
            # Update metadata
            report_metadata.delivered_to.extend(recipients)
            report_metadata.status = ReportStatus.DELIVERED
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to deliver report: {e}")
            return False
            
    async def _deliver_reports(
        self,
        reports: List[ReportMetadata],
        recipients: List[str]
    ):
        """Deliver multiple reports"""
        for report in reports:
            await self.deliver_report(report, recipients)
            
    # Custom Report Builder
    
    async def build_custom_report(
        self,
        query: str,
        columns: List[str],
        filters: Optional[Dict[str, Any]] = None,
        aggregations: Optional[Dict[str, str]] = None,
        charts: Optional[List[ChartConfig]] = None,
        output_format: ReportFormat = ReportFormat.PDF
    ) -> ReportMetadata:
        """Build a custom report using query builder"""
        # This would typically connect to a database
        # For now, we'll simulate with sample data
        
        # Generate sample data based on query
        data = pd.DataFrame({
            'date': pd.date_range('2024-01-01', periods=100),
            'sales': np.random.randint(100, 1000, 100),
            'customers': np.random.randint(10, 100, 100),
            'region': np.random.choice(['North', 'South', 'East', 'West'], 100),
            'product': np.random.choice(['A', 'B', 'C', 'D'], 100),
        })
        
        # Apply filters
        if filters:
            for column, value in filters.items():
                if column in data.columns:
                    if isinstance(value, list):
                        data = data[data[column].isin(value)]
                    else:
                        data = data[data[column] == value]
                        
        # Apply aggregations
        if aggregations:
            group_cols = [col for col in aggregations.keys() if aggregations[col] == 'group']
            agg_dict = {
                col: func for col, func in aggregations.items()
                if func != 'group' and col in data.columns
            }
            if group_cols and agg_dict:
                data = data.groupby(group_cols).agg(agg_dict).reset_index()
                
        # Select columns
        if columns:
            data = data[columns]
            
        # Create custom template
        custom_template = """
        <h1>Custom Report</h1>
        <p>Generated on: {{ generated_at }}</p>
        
        <h2>Data Summary</h2>
        <table>
            <thead>
                <tr>
                    {% for col in data[0].keys() %}
                    <th>{{ col }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in data %}
                <tr>
                    {% for value in row.values() %}
                    <td>{{ value }}</td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
        """
        
        # Generate report
        return await self.generate_report(
            data=data,
            output_format=output_format,
            custom_template=custom_template
        )
        
    # Cache Management
    
    def _generate_cache_key(
        self,
        template_id: str,
        parameters: Optional[Dict[str, Any]],
        output_format: ReportFormat
    ) -> str:
        """Generate cache key for report"""
        cache_data = {
            'template_id': template_id,
            'parameters': parameters or {},
            'format': output_format.value
        }
        cache_str = json.dumps(cache_data, sort_keys=True)
        return hashlib.md5(cache_str.encode()).hexdigest()
        
    async def _get_cached_report(self, cache_key: str) -> Optional[ReportMetadata]:
        """Get cached report if exists and not expired"""
        for report in self.report_history:
            if (
                report.cache_key == cache_key and
                report.status == ReportStatus.COMPLETED and
                report.expires_at and
                report.expires_at > datetime.utcnow()
            ):
                return report
        return None
        
    async def _cache_report(self, metadata: ReportMetadata):
        """Cache report metadata"""
        # In production, this would use Redis or similar
        pass
        
    # Report History and Cleanup
    
    async def get_report_history(
        self,
        limit: int = 100,
        offset: int = 0,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[ReportMetadata]:
        """Get report generation history"""
        history = sorted(
            self.report_history,
            key=lambda x: x.generated_at,
            reverse=True
        )
        
        # Apply filters
        if filters:
            if 'template_id' in filters:
                history = [r for r in history if r.template_id == filters['template_id']]
            if 'status' in filters:
                history = [r for r in history if r.status == filters['status']]
            if 'format' in filters:
                history = [r for r in history if r.format == filters['format']]
            if 'date_from' in filters:
                history = [r for r in history if r.generated_at >= filters['date_from']]
            if 'date_to' in filters:
                history = [r for r in history if r.generated_at <= filters['date_to']]
                
        return history[offset:offset + limit]
        
    async def _cleanup_old_reports(self, retention_days: int):
        """Clean up reports older than retention period"""
        cutoff_date = datetime.utcnow() - timedelta(days=retention_days)
        
        for report in self.report_history[:]:
            if report.generated_at < cutoff_date:
                # Delete file
                if report.file_path and os.path.exists(report.file_path):
                    os.remove(report.file_path)
                    
                # Remove from history
                self.report_history.remove(report)
                
    # Multi-language Support
    
    async def _get_translations(self, language: str) -> Translations:
        """Get translations for a language"""
        if language not in self.translations:
            try:
                translations = gettext.translation(
                    'reports',
                    localedir=str(self.locale_path),
                    languages=[language]
                )
                self.translations[language] = translations
            except:
                # Fallback to null translations
                self.translations[language] = gettext.NullTranslations()
                
        return self.translations[language]
        
    # Storage Management
    
    async def _save_templates(self):
        """Save templates to disk"""
        templates_file = self.storage_path / 'templates.json'
        templates_data = [
            {
                'id': t.id,
                'name': t.name,
                'description': t.description,
                'template_type': t.template_type,
                'template_content': t.template_content,
                'template_file': t.template_file,
                'data_source': t.data_source,
                'parameters': t.parameters,
                'charts': [
                    {
                        'chart_type': c.chart_type.value,
                        'data_columns': c.data_columns,
                        'x_label': c.x_label,
                        'y_label': c.y_label,
                        'title': c.title,
                        'colors': c.colors,
                        'width': c.width,
                        'height': c.height
                    }
                    for c in t.charts
                ],
                'styles': t.styles,
                'language': t.language,
                'created_at': t.created_at.isoformat(),
                'updated_at': t.updated_at.isoformat(),
                'tags': t.tags
            }
            for t in self.templates.values()
        ]
        
        async with aiofiles.open(templates_file, 'w') as f:
            await f.write(json.dumps(templates_data, indent=2))
            
    async def _load_templates(self):
        """Load templates from disk"""
        templates_file = self.storage_path / 'templates.json'
        if not templates_file.exists():
            return
            
        async with aiofiles.open(templates_file, 'r') as f:
            templates_data = json.loads(await f.read())
            
        for data in templates_data:
            charts = [
                ChartConfig(
                    chart_type=ChartType(c['chart_type']),
                    data_columns=c['data_columns'],
                    x_label=c.get('x_label'),
                    y_label=c.get('y_label'),
                    title=c.get('title'),
                    colors=c.get('colors'),
                    width=c.get('width', 8),
                    height=c.get('height', 6)
                )
                for c in data.get('charts', [])
            ]
            
            template = ReportTemplate(
                id=data['id'],
                name=data['name'],
                description=data.get('description', ''),
                template_type=data.get('template_type', 'jinja2'),
                template_content=data.get('template_content', ''),
                template_file=data.get('template_file'),
                data_source=data.get('data_source'),
                parameters=data.get('parameters', {}),
                charts=charts,
                styles=data.get('styles', {}),
                language=data.get('language', 'en'),
                created_at=datetime.fromisoformat(data['created_at']),
                updated_at=datetime.fromisoformat(data['updated_at']),
                tags=data.get('tags', [])
            )
            
            self.templates[template.id] = template
            
    async def _save_schedules(self):
        """Save schedules to disk"""
        schedules_file = self.storage_path / 'schedules.json'
        schedules_data = [
            {
                'id': s.id,
                'report_template_id': s.report_template_id,
                'cron_expression': s.cron_expression,
                'timezone': s.timezone,
                'enabled': s.enabled,
                'recipients': s.recipients,
                'output_formats': [f.value for f in s.output_formats],
                'parameters': s.parameters,
                'retention_days': s.retention_days,
                'created_at': s.created_at.isoformat(),
                'last_run': s.last_run.isoformat() if s.last_run else None,
                'next_run': s.next_run.isoformat() if s.next_run else None
            }
            for s in self.schedules.values()
        ]
        
        async with aiofiles.open(schedules_file, 'w') as f:
            await f.write(json.dumps(schedules_data, indent=2))
            
    async def _load_schedules(self):
        """Load schedules from disk"""
        schedules_file = self.storage_path / 'schedules.json'
        if not schedules_file.exists():
            return
            
        async with aiofiles.open(schedules_file, 'r') as f:
            schedules_data = json.loads(await f.read())
            
        for data in schedules_data:
            schedule = ReportSchedule(
                id=data['id'],
                report_template_id=data['report_template_id'],
                cron_expression=data['cron_expression'],
                timezone=data.get('timezone', 'UTC'),
                enabled=data.get('enabled', True),
                recipients=data.get('recipients', []),
                output_formats=[ReportFormat(f) for f in data.get('output_formats', ['pdf'])],
                parameters=data.get('parameters', {}),
                retention_days=data.get('retention_days', 30),
                created_at=datetime.fromisoformat(data['created_at']),
                last_run=datetime.fromisoformat(data['last_run']) if data.get('last_run') else None,
                next_run=datetime.fromisoformat(data['next_run']) if data.get('next_run') else None
            )
            
            self.schedules[schedule.id] = schedule
            
    async def _setup_scheduled_reports(self):
        """Set up all scheduled reports on startup"""
        for schedule in self.schedules.values():
            if schedule.enabled:
                await self._add_scheduled_job(schedule)
                
    async def _fetch_data(
        self,
        data_source: str,
        parameters: Optional[Dict[str, Any]] = None
    ) -> pd.DataFrame:
        """Fetch data from a data source"""
        # This would typically connect to databases, APIs, etc.
        # For now, return sample data
        
        import numpy as np
        
        return pd.DataFrame({
            'date': pd.date_range('2024-01-01', periods=30),
            'value': np.random.randint(100, 1000, 30),
            'category': np.random.choice(['A', 'B', 'C'], 30)
        })


# Example usage
if __name__ == "__main__":
    import asyncio
    
    async def main():
        # Initialize service
        report_service = ReportService(
            smtp_config={
                'host': 'smtp.gmail.com',
                'port': 587,
                'use_tls': True,
                'username': 'your-email@gmail.com',
                'password': 'your-password',
                'from_email': 'reports@example.com'
            }
        )
        
        await report_service.initialize()
        
        # Create a template
        template = await report_service.create_template(
            ReportTemplate(
                name="Sales Report",
                description="Monthly sales report with charts",
                template_content="""
                    <h1>Sales Report for {{ parameters.month }}</h1>
                    <p>Total Sales: {{ sum('sales') }}</p>
                    <table>
                        <tr>
                            <th>Date</th>
                            <th>Sales</th>
                            <th>Region</th>
                        </tr>
                        {% for row in data %}
                        <tr>
                            <td>{{ row.date }}</td>
                            <td>{{ row.sales }}</td>
                            <td>{{ row.region }}</td>
                        </tr>
                        {% endfor %}
                    </table>
                """,
                charts=[
                    ChartConfig(
                        chart_type=ChartType.BAR,
                        data_columns=['date', 'sales'],
                        title="Daily Sales",
                        x_label="Date",
                        y_label="Sales ($)"
                    )
                ],
                tags=["sales", "monthly"]
            )
        )
        
        # Generate report
        report = await report_service.generate_report(
            template_id=template.id,
            parameters={'month': 'January 2024'},
            output_format=ReportFormat.PDF
        )
        
        print(f"Report generated: {report.file_path}")
        
        # Create schedule
        schedule = await report_service.create_schedule(
            ReportSchedule(
                report_template_id=template.id,
                cron_expression="0 9 1 * *",  # First day of month at 9 AM
                recipients=["manager@example.com"],
                output_formats=[ReportFormat.PDF, ReportFormat.EXCEL],
                parameters={'month': 'Current Month'}
            )
        )
        
        print(f"Schedule created: Next run at {schedule.next_run}")
        
        # Generate custom report
        custom_report = await report_service.build_custom_report(
            query="SELECT * FROM sales",
            columns=['date', 'sales', 'region'],
            filters={'region': ['North', 'South']},
            aggregations={'region': 'group', 'sales': 'sum'},
            charts=[
                ChartConfig(
                    chart_type=ChartType.PIE,
                    data_columns=['region', 'sales'],
                    title="Sales by Region"
                )
            ]
        )
        
        print(f"Custom report generated: {custom_report.file_path}")
        
        await report_service.shutdown()
        
    asyncio.run(main())