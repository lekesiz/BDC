_('api_reports.label.reports_api_endpoints')
from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime, timedelta
from sqlalchemy import func
import os
import io
import csv
import json
from openpyxl import Workbook
from app.extensions import db
from app.models.user import User
from app.models.report import Report, ReportSchedule
from app.models.tenant import Tenant
from app.utils.logging import logger
from flask_babel import _, lazy_gettext as _l
reports_bp = Blueprint('reports', __name__)
DEMO_REPORT_TYPES = ['beneficiary', 'program', 'trainer', 'analytics',
    'performance']
DEMO_REPORT_NAMES = {'beneficiary': [_(
    'api_reports.label.beneficiary_progress_report'), _(
    'api_reports.label.monthly_beneficiary_summary'), _(
    'api_reports.label.beneficiary_test_results')], 'program': [_(
    'api_reports.label.program_performance_report'), _(
    'api_reports.label.program_enrollment_summary'), _(
    'api_reports.label.program_completion_analysis')], 'trainer': [_(
    'api_reports.label.trainer_activity_report'), _(
    'api_reports.label.trainer_performance_summary'), _(
    'api_reports.label.trainer_load_analysis')], 'analytics': [_(
    'api_reports.label.overall_analytics_report'), _(
    'api_reports.label.kpi_dashboard_summary'), _(
    'api_reports.label.trend_analysis_report')], 'performance': [_(
    'api_reports.label.performance_metrics_report'), _(
    'api_reports.label.test_score_analysis'), _(
    'api_reports.label.learning_outcomes_report')]}
DEMO_REPORT_DESCRIPTIONS = {'beneficiary': _(
    'api_reports.message.comprehensive_report_on_benefi'), 'program': _(
    'api_reports.message.detailed_analysis_of_program_p'), 'trainer': _(
    'api_reports.message.summary_of_trainer_activities'), 'analytics': _(
    'api_reports.message.overall_system_analytics_and_k'), 'performance': _
    ('api_reports.message.detailed_performance_metrics_a')}


@reports_bp.route('/reports/recent', methods=['GET'])
@jwt_required()
def get_recent_reports():
    _('api_reports.message.get_recent_reports_for_the_cur')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    query = Report.query.filter_by(created_by_id=user_id)
    if user.role in ['super_admin', 'tenant_admin']:
        query = Report.query.filter_by(tenant_id=user.tenant_id)
    recent_reports = query.order_by(Report.created_at.desc()).limit(10).all()
    return jsonify([report.to_dict() for report in recent_reports]), 200


@reports_bp.route('/reports/saved', methods=['GET'])
@jwt_required()
def get_saved_reports():
    _('api_reports.success.get_saved_reports_for_the_curr')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    query = Report.query.filter_by(status='completed')
    if user.role in ['super_admin', 'tenant_admin']:
        query = query.filter_by(tenant_id=user.tenant_id)
    else:
        query = query.filter_by(created_by_id=user_id)
    saved_reports = query.order_by(Report.updated_at.desc()).all()
    return jsonify([report.to_dict() for report in saved_reports]), 200


@reports_bp.route('/reports/scheduled', methods=['GET'])
@jwt_required()
def get_scheduled_reports():
    _('api_reports.message.get_scheduled_reports_for_the')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    query = ReportSchedule.query.join(Report).filter(ReportSchedule.
        is_active == True)
    if user.role in ['super_admin', 'tenant_admin']:
        query = query.filter(Report.tenant_id == user.tenant_id)
    else:
        query = query.filter(Report.created_by_id == user_id)
    scheduled_reports = query.all()
    return jsonify([schedule.to_dict() for schedule in scheduled_reports]), 200


@reports_bp.route('/reports', methods=['POST'])
@jwt_required()
def create_report():
    _('api_reports.message.create_a_new_report')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    data = request.get_json()
    try:
        report = Report(name=data['name'], description=data.get(
            'description'), type=data['type'], format=data.get('format',
            'pdf'), parameters=data.get('parameters', {}), created_by_id=
            user_id, tenant_id=user.tenant_id)
        db.session.add(report)
        db.session.commit()
        return jsonify(report.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/<int:report_id>', methods=['GET'])
@jwt_required()
def get_report(report_id):
    _('api_reports.message.get_a_specific_report')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    return jsonify(report.to_dict()), 200


@reports_bp.route('/reports/<int:report_id>', methods=['PUT'])
@jwt_required()
def update_report(report_id):
    """Update a report."""
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    data = request.get_json()
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    report.name = data.get('name', report.name)
    report.description = data.get('description', report.description)
    report.parameters = data.get('parameters', report.parameters)
    db.session.commit()
    return jsonify(report.to_dict()), 200


@reports_bp.route('/reports/<int:report_id>/run', methods=['POST'])
@jwt_required()
def run_report(report_id):
    _('api_reports.message.run_a_report_and_generate_outp')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    try:
        report.status = 'generating'
        db.session.commit()
        if report.type == 'beneficiary':
            data = generate_beneficiary_report(report.parameters)
        elif report.type == 'trainer':
            data = generate_trainer_report(report.parameters)
        elif report.type == 'program':
            data = generate_program_report(report.parameters)
        elif report.type == 'performance':
            data = generate_performance_report(report.parameters)
        else:
            data = generate_general_report(report.parameters)
        if report.format == 'pdf':
            file_path = generate_pdf_report(data, report)
        elif report.format == 'xlsx':
            file_path = generate_excel_report(data, report)
        else:
            file_path = generate_csv_report(data, report)
        report.status = 'completed'
        report.file_path = file_path
        report.file_size = os.path.getsize(file_path)
        report.last_generated = datetime.utcnow()
        report.run_count += 1
        db.session.commit()
        return jsonify(report.to_dict()), 200
    except Exception as e:
        report.status = 'failed'
        db.session.commit()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/<int:report_id>/download', methods=['GET'])
@jwt_required()
def download_report(report_id):
    _('api_reports.message.download_a_generated_report')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    if not report.file_path or not os.path.exists(report.file_path):
        return jsonify({'error': _(
            'api_reports.message.report_file_not_found')}), 404
    return send_file(report.file_path, as_attachment=True, download_name=
        f"{report.name}_{datetime.now().strftime('%Y%m%d')}.{report.format}")


@reports_bp.route('/reports/<int:report_id>/download-pdf', methods=['GET'])
@jwt_required()
def download_report_pdf(report_id):
    _('api_reports.message.download_a_report_as_pdf')
    from app.utils.pdf_generator import generate_report_pdf
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    try:
        if report.type == 'beneficiary':
            data = generate_beneficiary_report(report.parameters)
        elif report.type == 'trainer':
            data = generate_trainer_report(report.parameters)
        elif report.type == 'program':
            data = generate_program_report(report.parameters)
        elif report.type == 'performance':
            data = generate_performance_report(report.parameters)
        else:
            data = generate_general_report(report.parameters)
        template_data = {'title': report.name, 'description': report.
            description, 'generated_at': datetime.now().strftime(_(
            'analytics_report_generator.message.y_m_d_h_m_s_1')),
            'generated_by': f'{user.first_name} {user.last_name}', 'data':
            data, 'report_type': report.type}
        pdf_data = generate_report_pdf(template_data)
        return send_file(io.BytesIO(pdf_data), mimetype='application/pdf',
            as_attachment=True, download_name=
            f"{report.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def generate_beneficiary_report(parameters):
    _('api_reports.message.generate_beneficiary_report_da')
    from app.models.beneficiary import Beneficiary
    from app.models.test import TestSession
    query = Beneficiary.query
    if 'beneficiary_ids' in parameters:
        query = query.filter(Beneficiary.id.in_(parameters['beneficiary_ids']))
    beneficiaries = query.all()
    data = []
    for beneficiary in beneficiaries:
        test_scores = db.session.query(func.avg(TestSession.score)).filter(
            TestSession.beneficiary_id == beneficiary.id, TestSession.
            status == 'completed').scalar() or 0
        data.append({_('services_beneficiary_service.label.name_2'):
            f'{beneficiary.first_name} {beneficiary.last_name}', _(
            'i18n_translation_service.label.email'): beneficiary.email, _(
            'i18n_translation_service.label.status'): beneficiary.status, _
            ('analytics_real_time_dashboard.label.average_score'): round(
            test_scores, 2), _('api_reports.success.created'): beneficiary.
            created_at.strftime(_(
            'analytics_predictive_analytics.message.y_m_d')) if beneficiary
            .created_at else ''})
    return data


def generate_trainer_report(parameters):
    _('api_reports.message.generate_trainer_report_data')
    from app.models.beneficiary import Beneficiary
    query = User.query.filter_by(role='trainer')
    if 'trainer_ids' in parameters:
        query = query.filter(User.id.in_(parameters['trainer_ids']))
    trainers = query.all()
    data = []
    for trainer in trainers:
        beneficiary_count = Beneficiary.query.filter_by(trainer_id=trainer.id
            ).count()
        data.append({_('services_beneficiary_service.label.name_2'):
            f'{trainer.first_name} {trainer.last_name}', _(
            'i18n_translation_service.label.email'): trainer.email, _(
            'i18n_translation_service.label.beneficiaries'):
            beneficiary_count, _('api_reports.label.active_7'): _(
            'i18n_translation_service.label.yes') if trainer.is_active else
            'No', _('api_reports.label.last_login_1'): trainer.last_login.
            strftime(_('api_reports.message.y_m_d_h_m_1')) if trainer.
            last_login else ''})
    return data


def generate_program_report(parameters):
    _('api_reports.message.generate_program_report_data')
    from app.models.program import Program, ProgramEnrollment, Module, Session, SessionAttendance
    query = Program.query
    if 'program_ids' in parameters:
        query = query.filter(Program.id.in_(parameters['program_ids']))
    if 'status' in parameters:
        query = query.filter_by(status=parameters['status'])
    programs = query.all()
    data = []
    for program in programs:
        enrollment_count = ProgramEnrollment.query.filter_by(program_id=
            program.id).count()
        active_enrollments = ProgramEnrollment.query.filter_by(program_id=
            program.id, status='enrolled').count()
        completed_enrollments = ProgramEnrollment.query.filter_by(program_id
            =program.id, status='completed').count()
        sessions = Session.query.filter_by(program_id=program.id).all()
        total_attendance = 0
        total_possible = 0
        for session in sessions:
            attendances = SessionAttendance.query.filter_by(session_id=
                session.id).all()
            present_count = sum(1 for a in attendances if a.
                attendance_status == 'present')
            total_attendance += present_count
            total_possible += len(attendances)
        attendance_rate = (total_attendance / total_possible * 100 if 
            total_possible > 0 else 0)
        data.append({_(
            'reporting_report_builder_service.label.program_name'): program
            .name, _('api_reports.label.code'): program.code, _(
            'i18n_translation_service.label.status'): program.status, _(
            'api_reports.label.total_enrollments'): enrollment_count, _(
            'api_reports.label.active_7'): active_enrollments, _(
            'api_reports.success.completed_3'): completed_enrollments, _(
            'reporting_report_builder_service.label.completion_rate'): 
            f'{completed_enrollments / enrollment_count * 100:.1f}%' if 
            enrollment_count > 0 else '0%', _(
            'api_reports.label.attendance_rate_1'):
            f'{attendance_rate:.1f}%', _(
            'i18n_translation_service.label.start_date'): program.
            start_date.strftime(_(
            'analytics_predictive_analytics.message.y_m_d')) if program.
            start_date else '', _('i18n_translation_service.label.end_date'
            ): program.end_date.strftime(_(
            'analytics_predictive_analytics.message.y_m_d')) if program.
            end_date else ''})
    return data


def generate_performance_report(parameters):
    _('api_reports.message.generate_performance_report_da')
    from app.models.test import TestSession
    from app.models.beneficiary import Beneficiary
    start_date = parameters.get('start_date', datetime.now() - timedelta(
        days=30))
    end_date = parameters.get('end_date', datetime.now())
    sessions = TestSession.query.filter(TestSession.completed_at.between(
        start_date, end_date), TestSession.status == 'completed').join(
        Beneficiary).all()
    data = []
    for session in sessions:
        data.append({_('api_reports.label.date_2'): session.completed_at.
            strftime(_('api_reports.message.y_m_d_h_m_1')), _(
            'api_reports.label.beneficiary'):
            f'{session.beneficiary.first_name} {session.beneficiary.last_name}'
            , _('core_user_service_example.label.test'): session.test.title if
            session.test else _('analytics_data_export.label.unknown'), _(
            'reporting_report_builder_service.label.score'): session.score,
            _('testing_benchmark_runner.label.duration'): session.duration,
            _('i18n_translation_service.label.status'): session.status})
    return data


def generate_general_report(parameters):
    _('api_reports.message.generate_general_report_data')
    return []


def generate_excel_report(data, report):
    _('api_reports.message.generate_excel_report_file')
    wb = Workbook()
    ws = wb.active
    ws.title = report.name[:31]
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for row_num, row_data in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col, value=row_data.get(header, '')
                    )
    reports_dir = os.path.join('app', 'static', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    filename = (
        f"report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    file_path = os.path.join(reports_dir, filename)
    wb.save(file_path)
    return file_path


def generate_csv_report(data, report):
    _('api_reports.message.generate_csv_report_file')
    reports_dir = os.path.join('app', 'static', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    filename = (
        f"report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
    file_path = os.path.join(reports_dir, filename)
    if data:
        headers = list(data[0].keys())
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
    else:
        open(file_path, 'a').close()
    return file_path


def generate_pdf_report(data, report):
    _('api_reports.message.generate_pdf_report_file')
    from app.utils.pdf_generator import generate_report_pdf
    reports_dir = os.path.join('app', 'static', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    filename = (
        f"report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
    file_path = os.path.join(reports_dir, filename)
    template_data = {'title': report.name, 'description': report.
        description, 'generated_at': datetime.now().strftime(_(
        'analytics_report_generator.message.y_m_d_h_m_s_1')),
        'generated_by':
        f'{report.created_by.first_name} {report.created_by.last_name}',
        'data': data, 'report_type': report.type}
    pdf_data = generate_report_pdf(template_data)
    with open(file_path, 'wb') as pdf_file:
        pdf_file.write(pdf_data)
    return file_path


@reports_bp.route('/reports/demo', methods=['POST'])
@jwt_required()
def create_demo_reports():
    _('api_reports.message.create_demo_reports_for_testin')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    try:
        created_reports = []
        for report_type in DEMO_REPORT_TYPES:
            for i, name in enumerate(DEMO_REPORT_NAMES[report_type]):
                report = Report(name=name, description=
                    DEMO_REPORT_DESCRIPTIONS[report_type], type=report_type,
                    format=['pdf', 'xlsx', 'csv'][i % 3], status=
                    'completed', parameters={'date_range': _(
                    'api_reports.message.last_30_days'), 'filters': {
                    'active': True}}, created_by_id=user_id, tenant_id=user
                    .tenant_id, run_count=i + 1, last_generated=datetime.
                    now() - timedelta(days=i))
                if i % 5 == 0:
                    report.status = 'draft'
                elif i % 7 == 0:
                    report.status = 'generating'
                db.session.add(report)
                created_reports.append(report)
        for i, report in enumerate(created_reports[:5]):
            schedule = ReportSchedule(report_id=report.id, frequency=[
                'daily', 'weekly', 'monthly'][i % 3], schedule_time=
                datetime.now().time(), recipients=[_(
                'api_reports.message.admin_bdc_com'), f'user{i}@bdc.com'],
                recipients_count=2, is_active=i % 2 == 0, status='active' if
                i % 2 == 0 else 'paused', next_run=datetime.now() +
                timedelta(days=i), last_run=datetime.now() - timedelta(days
                =i) if i > 0 else None)
            db.session.add(schedule)
        db.session.commit()
        return jsonify({'message': _(
            'api_reports.success.demo_reports_created_successfu'), 'count':
            len(created_reports)}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/<int:report_id>/export', methods=['GET'])
@jwt_required()
def export_report(report_id):
    _('api_reports.validation.export_report_in_different_for')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    format = request.args.get('format', 'pdf')
    report = Report.query.get_or_404(report_id)
    if user.role not in ['super_admin', 'tenant_admin']:
        if report.created_by_id != user_id:
            return jsonify({'error': _(
                'programs_v2_session_routes.label.unauthorized_4')}), 403
    try:
        if report.type == 'beneficiary':
            data = generate_beneficiary_report(report.parameters)
        elif report.type == 'trainer':
            data = generate_trainer_report(report.parameters)
        elif report.type == 'program':
            data = generate_program_report(report.parameters)
        elif report.type == 'performance':
            data = generate_performance_report(report.parameters)
        else:
            data = generate_general_report(report.parameters)
        if format == 'pdf':
            file_path = generate_pdf_report(data, report)
        elif format == 'xlsx':
            file_path = generate_excel_report(data, report)
        else:
            file_path = generate_csv_report(data, report)
        return send_file(file_path, as_attachment=True, download_name=
            f"{report.name}_{datetime.now().strftime('%Y%m%d')}.{format}")
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/fields/<report_type>', methods=['GET'])
@jwt_required()
def get_report_fields(report_type):
    _('api_reports.message.get_available_fields_for_a_rep')
    fields_by_type = {'beneficiary': [{'id': 'name', 'name': _(
        'api_reports.label.full_name'), 'description': _(
        'api_reports.label.beneficiary_full_name')}, {'id': 'email', 'name':
        _('i18n_translation_service.label.email'), 'description': _(
        'api_reports.label.beneficiary_email_address')}, {'id': 'status',
        'name': _('i18n_translation_service.label.status'), 'description':
        _('api_reports.label.current_beneficiary_status')}, {'id':
        'trainer', 'name': _('api_reports.label.assigned_trainer'),
        'description': _(
        'api_reports.message.trainer_assigned_to_beneficiar')}, {'id':
        'created_date', 'name': _('api_reports.success.created_date_1'),
        'description': _('api_reports.message.date_beneficiary_was_added')},
        {'id': 'test_score', 'name': _(
        'api_reports.label.average_test_score'), 'description': _(
        'api_reports.message.average_score_across_all_tests')}, {'id':
        'progress', 'name': _('api_reports.label.progress'), 'description':
        _('api_reports.label.overall_progress_percentage')}, {'id':
        'department', 'name': _('api_reports.label.department'),
        'description': _('api_reports.label.department_classification')}, {
        'id': 'notes', 'name': _('i18n_translation_service.label.notes'),
        'description': _(
        'api_reports.message.additional_notes_and_comments')}], 'program':
        [{'id': 'name', 'name': _(
        'reporting_report_builder_service.label.program_name'),
        'description': _('api_reports.message.name_of_the_training_program'
        )}, {'id': 'code', 'name': _('api_reports.label.program_code'),
        'description': _('api_reports.label.unique_program_identifier')}, {
        'id': 'status', 'name': _('i18n_translation_service.label.status'),
        'description': _('api_reports.label.current_program_status')}, {
        'id': 'start_date', 'name': _(
        'i18n_translation_service.label.start_date'), 'description': _(
        'api_reports.label.program_start_date')}, {'id': 'end_date', 'name':
        _('i18n_translation_service.label.end_date'), 'description': _(
        'api_reports.label.program_end_date')}, {'id': 'enrollment_count',
        'name': _('api_reports.label.enrollment_count_1'), 'description': _
        ('api_reports.message.number_of_enrolled_beneficiari')}, {'id':
        'completion_rate', 'name': _(
        'reporting_report_builder_service.label.completion_rate'),
        'description': _('api_reports.label.program_completion_percentage')
        }, {'id': 'attendance_rate', 'name': _(
        'api_reports.label.attendance_rate_1'), 'description': _(
        'api_reports.label.average_attendance_rate')}, {'id': 'description',
        'name': _('i18n_translation_service.label.description'),
        'description': _('api_reports.label.program_description')}],
        'trainer': [{'id': 'name', 'name': _(
        'api_reports.label.trainer_name'), 'description': _(
        'api_reports.message.full_name_of_the_trainer')}, {'id': 'email',
        'name': _('i18n_translation_service.label.email'), 'description': _
        ('api_reports.label.trainer_email_address')}, {'id':
        'beneficiary_count', 'name': _(
        'api_reports.label.beneficiary_count_1'), 'description': _(
        'api_reports.message.number_of_assigned_beneficiari')}, {'id':
        'active_status', 'name': _('api_reports.label.active_status_1'),
        'description': _('api_reports.message.whether_trainer_is_active')},
        {'id': 'last_login', 'name': _('api_reports.label.last_login_1'),
        'description': _('api_reports.label.last_login_timestamp')}, {'id':
        'specialization', 'name': _('api_reports.label.specialization'),
        'description': _('api_reports.label.areas_of_expertise')}, {'id':
        'programs', 'name': _('api_reports.label.assigned_programs'),
        'description': _(
        'api_reports.message.programs_trainer_is_involved_i')}, {'id':
        'performance_rating', 'name': _(
        'api_reports.label.performance_rating'), 'description': _(
        'api_reports.label.average_performance_rating')}], 'analytics': [{
        'id': 'metric_name', 'name': _('api_reports.label.metric_name'),
        'description': _('api_reports.message.name_of_the_metric')}, {'id':
        'value', 'name': _('services_beneficiary_service.label.value'),
        'description': _('api_reports.label.metric_value')}, {'id':
        'change', 'name': _('api_reports.label.change'), 'description':
        'Change from previous period'}, {'id': 'trend', 'name': _(
        'api_reports.label.trend'), 'description': _(
        'api_reports.label.trend_direction')}, {'id': 'date', 'name': _(
        'api_reports.label.date_2'), 'description': _(
        'api_reports.message.date_of_the_metric')}, {'id': 'category',
        'name': _('api_reports.label.category'), 'description': _(
        'api_reports.label.metric_category')}], 'performance': [{'id':
        'beneficiary_name', 'name': _('api_reports.label.beneficiary_name'),
        'description': _('api_reports.message.full_name_of_the_beneficiary'
        )}, {'id': 'test_name', 'name': _('api_reports.label.test_name_1'),
        'description': _('api_reports.message.name_of_the_test')}, {'id':
        'score', 'name': _('reporting_report_builder_service.label.score'),
        'description': _('api_reports.label.test_score')}, {'id': 'date',
        'name': _('api_reports.label.date_2'), 'description': _(
        'api_reports.label.test_completion_date')}, {'id': 'duration',
        'name': _('testing_benchmark_runner.label.duration'), 'description':
        _('api_reports.success.time_taken_to_complete')}, {'id': 'status',
        'name': _('i18n_translation_service.label.status'), 'description':
        _('api_reports.label.test_completion_status')}, {'id': 'feedback',
        'name': _('api_reports.label.feedback'), 'description': _(
        'api_reports.message.test_feedback_and_comments')}, {'id':
        'improvement', 'name': _('api_reports.label.improvement'),
        'description': 'Improvement from previous test'}]}
    fields = fields_by_type.get(report_type, [])
    return jsonify(fields), 200


@reports_bp.route('/reports/filters/<report_type>', methods=['GET'])
@jwt_required()
def get_report_filters(report_type):
    _('api_reports.message.get_available_filters_for_a_re')
    filters_by_type = {'beneficiary': [{'id': 'status', 'name': _(
        'i18n_translation_service.label.status'), 'type': 'select',
        'options': [{'value': 'active', 'label': _(
        'api_reports.label.active_7')}, {'value': 'inactive', 'label': _(
        'api_reports.label.inactive_1')}, {'value': 'completed', 'label': _
        ('api_reports.success.completed_3')}, {'value': 'pending', 'label':
        _('api_reports.label.pending')}]}, {'id': 'trainer', 'name': _(
        'api_reports.label.trainer_1'), 'type': 'multiselect', 'options': [
        ]}, {'id': 'test_score_range', 'name': _(
        'api_reports.label.test_score_range'), 'type': 'range', 'options':
        {'min': 0, 'max': 100}}, {'id': 'created_date', 'name': _(
        'api_reports.success.created_date_1'), 'type': 'date', 'options': {
        }}], 'program': [{'id': 'status', 'name': _(
        'i18n_translation_service.label.status'), 'type': 'select',
        'options': [{'value': 'active', 'label': _(
        'api_reports.label.active_7')}, {'value': 'completed', 'label': _(
        'api_reports.success.completed_3')}, {'value': 'upcoming', 'label':
        _('api_reports.label.upcoming')}, {'value': 'cancelled', 'label': _
        ('api_reports.label.cancelled')}]}, {'id': 'enrollment_range',
        'name': _('api_reports.label.enrollment_count_1'), 'type': 'range',
        'options': {'min': 0, 'max': 100}}, {'id': 'date_range', 'name': _(
        'api_reports.label.date_range_2'), 'type': 'date', 'options': {}}],
        'trainer': [{'id': 'active_status', 'name': _(
        'api_reports.label.active_status_1'), 'type': 'select', 'options':
        [{'value': 'active', 'label': _('api_reports.label.active_7')}, {
        'value': 'inactive', 'label': _('api_reports.label.inactive_1')}]},
        {'id': 'beneficiary_count', 'name': _(
        'api_reports.label.beneficiary_count_1'), 'type': 'range',
        'options': {'min': 0, 'max': 50}}, {'id': 'programs', 'name': _(
        'i18n_translation_service.label.programs'), 'type': 'multiselect',
        'options': []}], 'analytics': [{'id': 'metric_category', 'name': _(
        'api_reports.label.metric_category_1'), 'type': 'multiselect',
        'options': [{'value': 'engagement', 'label': _(
        'api_reports.label.engagement')}, {'value': 'performance', 'label':
        _('analytics_analytics_orchestrator.label.performance_1')}, {
        'value': 'completion', 'label': _('api_reports.label.completion')},
        {'value': 'satisfaction', 'label': _(
        'api_reports.label.satisfaction')}]}, {'id': 'date_range', 'name':
        _('api_reports.label.date_range_2'), 'type': 'date', 'options': {}}
        ], 'performance': [{'id': 'test_name', 'name': _(
        'api_reports.label.test_name_1'), 'type': 'multiselect', 'options':
        []}, {'id': 'score_range', 'name': _(
        'api_reports.label.score_range'), 'type': 'range', 'options': {
        'min': 0, 'max': 100}}, {'id': 'status', 'name': _(
        'i18n_translation_service.label.status'), 'type': 'select',
        'options': [{'value': 'completed', 'label': _(
        'api_reports.success.completed_3')}, {'value': 'in_progress',
        'label': _('api_reports.label.in_progress')}, {'value': 'abandoned',
        'label': _('api_reports.label.abandoned')}]}, {'id': 'date_range',
        'name': _('api_reports.label.date_range_2'), 'type': 'date',
        'options': {}}]}
    filters = filters_by_type.get(report_type, [])
    if report_type == 'beneficiary' and filters:
        trainers = User.query.filter_by(role='trainer').all()
        for filter in filters:
            if filter['id'] == 'trainer':
                filter['options'] = [{'value': str(t.id), 'label':
                    f'{t.first_name} {t.last_name}'} for t in trainers]
    return jsonify(filters), 200


@reports_bp.route('/reports/preview', methods=['POST'])
@jwt_required()
def preview_report():
    _('api_reports.message.generate_a_preview_of_the_repo')
    data = request.get_json()
    try:
        preview_data = {'sections': []}
        if data['type'] == 'beneficiary':
            preview_data['sections'].append({'title': _(
                'ai_content_recommendations.label.summary_1'), 'type':
                'summary', 'metrics': [{'id': 'total', 'name': _(
                'api_reports.label.total_beneficiaries'), 'value': '125',
                'change': 12}, {'id': 'active', 'name': _(
                'api_reports.label.active_7'), 'value': '98', 'change': 5},
                {'id': 'avg_score', 'name': _(
                'analytics_real_time_dashboard.label.average_score'),
                'value': _('api_reports.message.82_5'), 'change': 3}]})
            preview_data['sections'].append({'title': _(
                'api_reports.label.beneficiary_details'), 'type': 'table',
                'columns': [{'id': 'name', 'name': _(
                'services_beneficiary_service.label.name_2')}, {'id':
                'email', 'name': _('i18n_translation_service.label.email')},
                {'id': 'status', 'name': _(
                'i18n_translation_service.label.status')}, {'id':
                'test_score', 'name': _('api_reports.label.avg_score')}, {
                'id': 'trainer', 'name': _('api_reports.label.trainer_1')}],
                'data': [{'name': _('api_reports.label.john_doe'), 'email':
                _('api_reports.message.john_doe_example_com'), 'status': _(
                'api_reports.label.active_7'), 'test_score': _(
                'api_reports.message.85'), 'trainer': _(
                'api_reports.label.sarah_johnson')}, {'name': _(
                'api_reports.label.jane_smith'), 'email': _(
                'api_reports.message.jane_smith_example_com'), 'status': _(
                'api_reports.label.active_7'), 'test_score': _(
                'api_reports.message.92'), 'trainer': _(
                'api_reports.label.mike_wilson')}]})
        elif data['type'] == 'program':
            preview_data['sections'].append({'title': _(
                'api_reports.label.program_enrollment_trends'), 'type':
                'chart', _('api_reports.message.charttype'): 'line', 'data':
                {}})
            preview_data['sections'].append({'title': _(
                'api_reports.label.program_details'), 'type': 'table',
                'columns': [{'id': 'name', 'name': _(
                'api_reports.label.program')}, {'id': 'enrollment', 'name':
                _('api_reports.label.enrollments')}, {'id': 'completion',
                'name': _(
                'reporting_report_builder_service.label.completion_rate')},
                {'id': 'attendance', 'name': _(
                'reporting_report_builder_service.label.attendance')}],
                'data': [{'name': _('api_reports.label.leadership_training'
                ), 'enrollment': '25', 'completion': _(
                'api_reports.message.88'), 'attendance': _(
                'api_reports.message.94')}, {'name': _(
                'programs_v2_util_routes.label.technical_skills'),
                'enrollment': '40', 'completion': _(
                'api_reports.message.76'), 'attendance': _(
                'api_reports.message.89')}]})
        return jsonify(preview_data), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@reports_bp.route('/reports/generate', methods=['POST'])
@jwt_required()
def generate_report():
    _('api_reports.validation.generate_a_report_in_the_reque')
    user_id = get_jwt_identity()
    user = User.query.get(user_id)
    data = request.get_json()
    try:
        report = Report(name=data.get('name', _(
            'api_reports.label.generated_report')), description=data.get(
            'description', ''), type=data['type'], format=data.get('format',
            'pdf'), parameters=data, created_by_id=user_id, tenant_id=user.
            tenant_id, status='generating')
        db.session.add(report)
        db.session.commit()
        if report.type == 'beneficiary':
            report_data = generate_beneficiary_report(data)
        elif report.type == 'trainer':
            report_data = generate_trainer_report(data)
        elif report.type == 'program':
            report_data = generate_program_report(data)
        elif report.type == 'performance':
            report_data = generate_performance_report(data)
        else:
            report_data = generate_general_report(data)
        if data.get('format') == 'pdf':
            file_path = generate_pdf_report(report_data, report)
        elif data.get('format') == 'xlsx':
            file_path = generate_excel_report(report_data, report)
        else:
            file_path = generate_csv_report(report_data, report)
        report.status = 'completed'
        report.file_path = file_path
        report.file_size = os.path.getsize(file_path)
        report.last_generated = datetime.now()
        db.session.commit()
        return send_file(file_path, as_attachment=True, download_name=
            f"{report.name}_{datetime.now().strftime('%Y%m%d')}.{report.format}"
            )
    except Exception as e:
        if 'report' in locals():
            report.status = 'failed'
            db.session.commit()
        return jsonify({'error': str(e)}), 500
